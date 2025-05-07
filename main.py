import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def scrape_fundamentus_fiis():
    url = "https://www.fundamentus.com.br/fii_resultado.php"
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
    }

    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()

        soup = BeautifulSoup(response.text, 'html.parser')
        table = soup.find('table', {'id': 'tabelaResultado'})

        # Extrair cabeçalhos
        headers = [th.get_text(strip=True) for th in table.find('tr').find_all('th')]

        # Extrair dados das linhas
        data = []
        for row in table.find_all('tr')[1:]:
            row_data = [td.get_text(strip=True) for td in row.find_all('td')]
            data.append(row_data)

        # Criar DataFrame
        df = pd.DataFrame(data, columns=headers)

        # Renomear colunas para consistência
        df = df.rename(columns={
            'Papel': 'Papel',
            'Segmento': 'Segmento',
            'Cotação': 'Cotacao',
            'FFO Yield': 'FFO Yield',
            'Dividend Yield': 'Dividend Yield',
            'P/VP': 'P/V',
            'Valor de Mercado': 'Valor de Mercado',
            'Liquidez': 'Liquidez',
            'Qtd de imóveis': 'Qtd Imoveis',
            'Preço do m2': 'Preco m2',
            'Aluguel por m2': 'Aluguel m2',
            'Cap Rate': 'Cap Rate',
            'Vacância Média': 'Vacancia Media'
        })

        return df

    except Exception as e:
        print(f"Erro ao acessar o site: {e}")
        return None


def clean_and_convert_data(df):
    # Converter valores numéricos
    df['Dividend Yield'] = df['Dividend Yield'].str.replace('%', '').str.replace('.', '').str.replace(',', '.').astype(
        float) / 100
    df['P/V'] = df['P/V'].str.replace(',', '.').astype(float)

    # Converter Valor de Mercado e Liquidez
    df['Valor de Mercado'] = df['Valor de Mercado'].apply(
        lambda x: float(x.replace('.', '')) if isinstance(x, str) else x)
    df['Liquidez'] = df['Liquidez'].apply(lambda x: float(x.replace('.', '')) if isinstance(x, str) else x)

    return df


def apply_filters(df):
    # Aplicar filtros
    filtered_df = df[
        (df['Dividend Yield'] > 0.10) &
        (df['Dividend Yield'] < 0.16) &
        (df['P/V'] > 0.6) &
        (df['P/V'] < 0.95) &
        (df['Liquidez'] > 1000000) &
        (df['Valor de Mercado'] > 1000000000)
        ].copy()

    return filtered_df


def calculate_score(row):
    score = 0

    # Dividend Yield (15% é o ideal)
    if row['Dividend Yield'] >= 0.15:
        score += 2
    elif row['Dividend Yield'] >= 0.13:
        score += 1

    # P/V (quanto menor melhor)
    if row['P/V'] <= 0.80:
        score += 2
    elif row['P/V'] <= 0.85:
        score += 1

    # Liquidez (quanto maior melhor)
    if row['Liquidez'] >= 5000000:
        score += 2
    elif row['Liquidez'] >= 2000000:
        score += 1

    # Valor de Mercado (quanto maior melhor)
    if row['Valor de Mercado'] >= 2000000000:
        score += 2
    elif row['Valor de Mercado'] >= 1500000000:
        score += 1

    return score


def style_excel_output(writer, df):
    # Acessar o workbook e worksheet
    workbook = writer.book
    worksheet = workbook['Sheet1']

    # Definir estilos
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    align_center = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Aplicar estilos ao cabeçalho
    for col in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = thin_border

        # Ajustar largura das colunas
        column_letter = get_column_letter(col)
        worksheet.column_dimensions[column_letter].width = 15

    # Aplicar estilos aos dados
    for row in range(2, len(df) + 2):
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

    # Formatar porcentagens
    for row in range(2, len(df) + 2):
        cell = worksheet.cell(row=row, column=4)  # Dividend Yield
        cell.number_format = '0.00%'

    # Congelar painel
    worksheet.freeze_panes = 'A2'


def main():
    print("Iniciando raspagem de dados do Fundamentus...")
    df = scrape_fundamentus_fiis()

    if df is not None:
        print("Dados obtidos com sucesso! Processando...")
        df = clean_and_convert_data(df)
        filtered_df = apply_filters(df)

        # Calcular nota para cada fundo
        filtered_df['Nota'] = filtered_df.apply(calculate_score, axis=1)

        # Selecionar e ordenar colunas
        final_df = filtered_df[['Papel', 'Segmento', 'Dividend Yield', 'P/V', 'Valor de Mercado', 'Liquidez', 'Nota']]
        final_df = final_df.sort_values(by='Nota', ascending=False)

        # Salvar em Excel
        output_file = "fundos_imobiliarios_filtrados.xlsx"
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            final_df.to_excel(writer, index=False, sheet_name='Sheet1')
            style_excel_output(writer, final_df)

        print(f"Processo concluído! Arquivo salvo como: {output_file}")
        print(f"Total de fundos encontrados: {len(final_df)}")
    else:
        print("Não foi possível obter os dados do site.")


if __name__ == "__main__":
    main()