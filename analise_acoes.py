import os
import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def verificar_arquivo_existente(nome_arquivo):
    """Verifica se o arquivo já existe e pergunta ao usuário se deseja substituí-lo"""
    if os.path.exists(nome_arquivo):
        print(f"\nAVISO: O arquivo '{nome_arquivo}' já existe no diretório.")
        # substituir = input("Deseja substituí-lo? (S/N): ").strip().upper()
        # return substituir == 'S'
    return True


def scrape_fundamentus_acoes():
    url = "https://www.fundamentus.com.br/resultado.php"
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
    }

    try:
        response = requests.get(url, headers=headers, timeout=10)
        response.raise_for_status()

        soup = BeautifulSoup(response.text, 'html.parser')
        table = soup.find('table', {'id': 'resultado'})

        if not table:
            print("Erro: Não foi possível encontrar a tabela de dados no site.")
            return None

        # Extrair cabeçalhos
        headers = [th.get_text(strip=True) for th in table.find('tr').find_all('th')]

        # Verificar se temos todas as colunas necessárias
        required_columns = ['Papel', 'P/L', 'P/VP', 'Div.Yield', 'ROE', 'Liq.2meses', 'Dív.Brut/ Patrim.', 'Cresc. Rec.5a']
        if not all(col in headers for col in required_columns):
            print("Algumas colunas necessárias não foram encontradas na tabela")
            print(f"Colunas encontradas: {headers}")
            return None

        # Extrair dados das linhas
        data = []
        for row in table.find_all('tr')[1:]:
            row_data = [td.get_text(strip=True) for td in row.find_all('td')]
            data.append(row_data)

        # Criar DataFrame
        df = pd.DataFrame(data, columns=headers)

        return df

    except requests.exceptions.RequestException as e:
        print(f"Erro ao acessar o site: {e}")
        return None
    except Exception as e:
        print(f"Erro inesperado: {e}")
        return None


def clean_and_convert_data(df):
    try:
        # Verificar se as colunas necessárias existem
        required_columns = ['P/L', 'P/VP', 'Div.Yield', 'ROE', 'Liq.2meses', 'Cresc. Rec.5a']
        for col in required_columns:
            if col not in df.columns:
                print(f"Coluna não encontrada: {col}")
                return None

        # Encontrar a coluna Dív.Brut/ Patrim. (pode ter nomes diferentes)
        div_patrim_col = None
        for col in df.columns:
            if 'Dív.Brut/ Patrim.' in col:
                div_patrim_col = col
                break

        if not div_patrim_col:
            print("Coluna 'Dív.Brut/ Patrim.' não encontrada")
            return None

        # Converter valores numéricos
        df['P/L'] = df['P/L'].str.replace('.', '').str.replace(',', '.').astype(float)
        df['P/VP'] = df['P/VP'].str.replace('.', '').str.replace(',', '.').astype(float)
        df['Div.Yield'] = df['Div.Yield'].str.replace('%', '').str.replace('.', '').str.replace(',', '.').astype(float) / 100
        df['ROE'] = df['ROE'].str.replace('%', '').str.replace('.', '').str.replace(',', '.').astype(float) / 100
        df['Liq.2meses'] = df['Liq.2meses'].str.replace('.', '').str.replace(',', '.').astype(float)
        df[div_patrim_col] = df[div_patrim_col].str.replace('.', '').str.replace(',', '.').astype(float)
        df['Cresc. Rec.5a'] = df['Cresc. Rec.5a'].str.replace('%', '').str.replace('.', '').str.replace(',',
                                                                                                        '.').astype(
            float) / 100

        # Renomear colunas para nomes consistentes
        df = df.rename(columns={
            'Div.Yield': 'Div.Yield',
            'Liq.2meses': 'Ltg.2meses',
            div_patrim_col: 'Dív.Brut/ Patrim.',
            'Cresc. Rec.5a': 'Cresc.Rec.5a'
        })

        return df
    except Exception as e:
        print(f"Erro ao converter dados: {e}")
        return None


def apply_filters(df):
    try:
        # Encontrar o nome exato da coluna
        div_patrim_col = [col for col in df.columns if 'Dív.Brut/ Patrim.' in col][0]

        # Aplicar filtros conforme especificado
        filtered_df = df[
            (df['P/L'] > 3.5) &
            (df['P/L'] < 12) &
            (df['P/VP'] > 0.5) &
            (df['P/VP'] < 1.1) &
            (df['ROE'] > 0.14) &
            (df['ROE'] < 0.40) &
            (df['Div.Yield'] > 0.07) &
            (df['Div.Yield'] < 0.20) &
            (df['Cresc.Rec.5a'] > 0.10) &
            (df[div_patrim_col] < 2) &
            (df['Ltg.2meses'] > 1000000)
            ].copy()

        return filtered_df
    except Exception as e:
        print(f"Erro ao aplicar filtros: {e}")
        return None


def calculate_score(row):
    try:
        score = 0

        # P/L (quanto menor melhor)
        if row['P/L'] <= 5:
            score += 2
        elif row['P/L'] <= 7:
            score += 1

        # P/VP (quanto menor melhor)
        if row['P/VP'] <= 0.7:
            score += 2
        elif row['P/VP'] <= 0.9:
            score += 1

        # Dividend Yield (quanto maior melhor)
        if row['Div.Yield'] >= 0.12:
            score += 2
        elif row['Div.Yield'] >= 0.09:
            score += 1

        # ROE (quanto maior melhor)
        if row['ROE'] >= 0.20:
            score += 2
        elif row['ROE'] >= 0.17:
            score += 1

        # Crescimento Receita (quanto maior melhor)
        if row['Cresc.Rec.5a'] >= 0.20:
            score += 2
        elif row['Cresc.Rec.5a'] >= 0.15:
            score += 1

        # Dívida Bruta/Patrimônio (quanto menor melhor)
        div_patrim_col = [col for col in row.index if 'Dív.Brut/ Patrim.' in col][0]
        if row[div_patrim_col] <= 0.5:
            score += 2
        elif row[div_patrim_col] <= 1:
            score += 1

        # Liquidez (quanto maior melhor)
        if row['Ltg.2meses'] >= 50000000:
            score += 2
        elif row['Ltg.2meses'] >= 10000000:
            score += 1

        return min(score, 12)  # Nota máxima de 12 pontos
    except Exception as e:
        print(f"Erro ao calcular score: {e}")
        return 0


def style_excel_output(writer, df):
    try:
        workbook = writer.book
        worksheet = workbook['Sheet1']

        # Definir estilos
        header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        align_center = Alignment(horizontal='center', vertical='center')
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        # Aplicar estilos ao cabeçalho
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center
            cell.border = thin_border

            # Ajustar largura das colunas
            column_letter = get_column_letter(col)
            worksheet.column_dimensions[column_letter].width = 15

        # Aplicar estilos aos dados
        for row in range(2, len(df) + 2):
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

        # Formatar porcentagens
        for col_name in ['Div.Yield', 'ROE', 'Cresc.Rec.5a']:
            col_idx = df.columns.get_loc(col_name) + 1
            for row in range(2, len(df) + 2):
                cell = worksheet.cell(row=row, column=col_idx)
                cell.number_format = '0.00%'

        # Formatar números decimais
        for col_name in ['P/L', 'P/VP']:
            col_idx = df.columns.get_loc(col_name) + 1
            for row in range(2, len(df) + 2):
                cell = worksheet.cell(row=row, column=col_idx)
                cell.number_format = '0.00'

        # Formatar Dív.Brut/ Patrim. (pode ter nomes diferentes)
        div_patrim_col = [col for col in df.columns if 'Dív.Brut/ Patrim.' in col][0]
        div_patrim_col_idx = df.columns.get_loc(div_patrim_col) + 1
        for row in range(2, len(df) + 2):
            cell = worksheet.cell(row=row, column=div_patrim_col_idx)
            cell.number_format = '0.00'

        # Formatar Liquidez como número inteiro
        ltg_col = df.columns.get_loc('Ltg.2meses') + 1
        for row in range(2, len(df) + 2):
            cell = worksheet.cell(row=row, column=ltg_col)
            cell.number_format = '#,##0'

        # Adicionar formatação condicional para a coluna Nota
        nota_col = df.columns.get_loc('Nota') + 1
        for row in range(2, len(df) + 2):
            cell = worksheet.cell(row=row, column=nota_col)
            if cell.value >= 8:
                cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            elif cell.value >= 5:
                cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            else:
                cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

        # Congelar painel
        worksheet.freeze_panes = 'A2'
    except Exception as e:
        print(f"Erro ao formatar arquivo Excel: {e}")


def main():
    print("\n=== Análise de Ações Fundamentus ===")
    print("Fonte: https://www.fundamentus.com.br/resultado.php")

    output_file = "acoes_filtradas_fundamentus.xlsx"

    # Verificar se o arquivo já existe
    if not verificar_arquivo_existente(output_file):
        print("Operação cancelada pelo usuário.")
        return

    print("\nIniciando raspagem de dados do Fundamentus...")
    df = scrape_fundamentus_acoes()

    if df is not None:
        print("Dados obtidos com sucesso! Processando...")
        df = clean_and_convert_data(df)

        if df is not None:
            filtered_df = apply_filters(df)

            if filtered_df is not None:
                if len(filtered_df) > 0:
                    # Calcular nota para cada ação
                    filtered_df['Nota'] = filtered_df.apply(calculate_score, axis=1)

                    # Selecionar e ordenar colunas
                    final_df = filtered_df[['Papel', 'P/L', 'P/VP', 'Div.Yield', 'ROE',
                                            'Ltg.2meses', 'Dív.Brut/ Patrim.', 'Cresc.Rec.5a', 'Nota']]
                    final_df = final_df.sort_values(by=['Nota', 'Div.Yield'], ascending=[False, False])

                    try:
                        # Salvar em Excel
                        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                            final_df.to_excel(writer, index=False, sheet_name='Sheet1')
                            style_excel_output(writer, final_df)

                        print(f"\nProcesso concluído com sucesso!")
                        print(f"Arquivo salvo como: {os.path.abspath(output_file)}")
                        print(f"Total de ações encontradas: {len(final_df)}")

                        if len(final_df) > 0:
                            print("\nTop 10 ações:")
                            print(final_df.head(10).to_string(index=False))
                    except Exception as e:
                        print(f"\nErro ao salvar o arquivo: {e}")
                else:
                    print("\nNenhuma ação atende aos critérios de filtro especificados.")
            else:
                print("\nErro ao filtrar os dados.")
        else:
            print("\nErro ao processar os dados.")
    else:
        print("\nNão foi possível obter os dados do site.")


if __name__ == "__main__":
    main()
