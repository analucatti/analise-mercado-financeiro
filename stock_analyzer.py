"""
Stock Analysis Scraper for Fundamentus
Scrapes and analyzes Brazilian stock market data with fundamental analysis metrics.
"""

import json
import logging
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from enum import Enum
from pathlib import Path
from typing import Optional, Dict, List

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


# ============= Configuration =============
@dataclass
class ScraperConfig:
    """Configuration for the stock scraper."""
    base_url: str = "https://www.fundamentus.com.br/resultado.php"
    details_url: str = "https://www.fundamentus.com.br/detalhes.php"
    timeout: int = 30
    user_agent: str = (
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 '
        '(KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
    )
    output_filename: str = "acoes_filtradas_fundamentus.xlsx"
    cache_filename: str = "setor_cache.json"
    max_workers: int = 5  # For parallel sector fetching
    rate_limit_delay: float = 0.5  # Delay between requests in seconds
    top_stocks_limit: int = 30
    top_per_sector: int = 5


@dataclass
class FilterCriteria:
    """Criteria for filtering stocks."""
    min_pl: float = 3.0  # P/L (Price/Earnings)
    max_pl: float = 12.0
    min_pvp: float = 0.5  # P/VP (Price/Book Value)
    max_pvp: float = 1.1
    min_roe: float = 0.14  # Return on Equity
    max_roe: float = 0.50
    min_dividend_yield: float = 0.07
    max_dividend_yield: float = 0.25
    min_revenue_growth: float = 0.10  # 5-year revenue growth
    max_debt_equity: float = 2.0  # Debt/Equity ratio
    min_liquidity: float = 1_000_000  # 2-month liquidity


@dataclass
class ScoreWeights:
    """Scoring weights and thresholds for stock evaluation."""
    # P/L thresholds (lower is better)
    pl_excellent: float = 5.0
    pl_good: float = 7.0

    # P/VP thresholds (lower is better)
    pvp_excellent: float = 0.7
    pvp_good: float = 0.9

    # Dividend Yield thresholds (higher is better)
    div_yield_excellent: float = 0.12
    div_yield_good: float = 0.09

    # ROE thresholds (higher is better)
    roe_excellent: float = 0.20
    roe_good: float = 0.17

    # Revenue Growth thresholds (higher is better)
    growth_excellent: float = 0.20
    growth_good: float = 0.15

    # Debt/Equity thresholds (lower is better)
    debt_excellent: float = 0.5
    debt_good: float = 1.0

    # Liquidity thresholds (higher is better)
    liquidity_excellent: float = 50_000_000
    liquidity_good: float = 10_000_000


class ColorScheme(Enum):
    """Excel color schemes for formatting."""
    HEADER_COLOR = '4F81BD'
    HEADER_TEXT = 'FFFFFF'
    HIGH_SCORE = 'C6EFCE'
    MEDIUM_SCORE = 'FFEB9C'
    LOW_SCORE = 'FFC7CE'


# ============= Column Mappings =============
COLUMN_MAPPINGS = {
    'Div.Yield': 'Div.Yield',
    'Liq.2meses': 'Ltg.2meses',
    'Dív.Brut/ Patrim.': 'Dív.Brut/Patrim',
    'Cresc. Rec.5a': 'Cresc.Rec.5a'
}

REQUIRED_COLUMNS = [
    'Papel', 'P/L', 'P/VP', 'Div.Yield', 'ROE',
    'Liq.2meses', 'Dív.Brut/ Patrim.', 'Cresc. Rec.5a'
]

FINAL_COLUMNS = [
    'Papel', 'Setor', 'P/L', 'P/VP', 'Div.Yield',
    'ROE', 'Ltg.2meses', 'Dív.Brut/Patrim', 'Cresc.Rec.5a', 'Nota'
]


# ============= Exceptions =============
class StockScraperError(Exception):
    """Base exception for stock scraper."""
    pass


class DataFetchError(StockScraperError):
    """Exception raised when data fetching fails."""
    pass


class DataProcessingError(StockScraperError):
    """Exception raised when data processing fails."""
    pass


# ============= Cache Management =============
class SectorCache:
    """Manages sector information caching."""

    def __init__(self, cache_file: Path):
        self.cache_file = cache_file
        self._cache: Dict[str, str] = {}
        self._lock = threading.Lock()
        self._modified = False
        self.load()

    def load(self) -> None:
        """Load cache from file."""
        if self.cache_file.exists():
            try:
                with open(self.cache_file, 'r', encoding='utf-8') as f:
                    self._cache = json.load(f)
                logger.info(f"Loaded {len(self._cache)} sectors from cache")
            except Exception as e:
                logger.warning(f"Could not load cache: {e}")
                self._cache = {}

    def save(self) -> None:
        """Save cache to file if modified."""
        if self._modified:
            try:
                with self._lock:
                    with open(self.cache_file, 'w', encoding='utf-8') as f:
                        json.dump(self._cache, f, ensure_ascii=False, indent=2)
                logger.info(f"Saved {len(self._cache)} sectors to cache")
                self._modified = False
            except Exception as e:
                logger.error(f"Could not save cache: {e}")

    def get(self, ticker: str) -> Optional[str]:
        """Get sector from cache."""
        with self._lock:
            return self._cache.get(ticker)

    def set(self, ticker: str, sector: str) -> None:
        """Add sector to cache."""
        with self._lock:
            if ticker not in self._cache or self._cache[ticker] != sector:
                self._cache[ticker] = sector
                self._modified = True

    def contains(self, ticker: str) -> bool:
        """Check if ticker is in cache."""
        with self._lock:
            return ticker in self._cache

    @property
    def size(self) -> int:
        """Get cache size."""
        with self._lock:
            return len(self._cache)

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.save()


# ============= Data Fetching =============
class FundamentusScraper:
    """Scraper for Fundamentus stock data."""

    def __init__(self, config: ScraperConfig):
        self.config = config
        self.session = self._create_session()
        self._last_request_time = 0
        self._request_lock = threading.Lock()

    def _create_session(self) -> requests.Session:
        """Create a configured requests session."""
        session = requests.Session()
        session.headers.update({'User-Agent': self.config.user_agent})
        return session

    def _rate_limit(self) -> None:
        """Implement rate limiting between requests."""
        with self._request_lock:
            current_time = time.time()
            time_since_last = current_time - self._last_request_time
            if time_since_last < self.config.rate_limit_delay:
                time.sleep(self.config.rate_limit_delay - time_since_last)
            self._last_request_time = time.time()

    def fetch_stock_list(self) -> pd.DataFrame:
        """
        Fetch the main stock list from Fundamentus.

        Returns:
            DataFrame with raw stock data

        Raises:
            DataFetchError: If data cannot be fetched or parsed
        """
        try:
            logger.info(f"Fetching stock list from {self.config.base_url}")
            response = self.session.get(
                self.config.base_url,
                timeout=self.config.timeout
            )
            response.raise_for_status()

            soup = BeautifulSoup(response.text, 'html.parser')
            table = soup.find('table', {'id': 'resultado'})

            if not table:
                raise DataFetchError("Table 'resultado' not found on page")

            df = self._parse_table(table)
            logger.info(f"Successfully fetched {len(df)} stocks")
            return df

        except requests.RequestException as e:
            logger.error(f"Request failed: {e}")
            raise DataFetchError(f"Failed to fetch data: {e}") from e
        except Exception as e:
            logger.error(f"Unexpected error during data fetch: {e}")
            raise DataFetchError(f"Unexpected error: {e}") from e

    def _parse_table(self, table) -> pd.DataFrame:
        """Parse HTML table into DataFrame."""
        headers = [th.get_text(strip=True) for th in table.find('tr').find_all('th')]

        data = []
        for row in table.find_all('tr')[1:]:
            row_data = [td.get_text(strip=True) for td in row.find_all('td')]
            if row_data:  # Skip empty rows
                data.append(row_data)

        if not data:
            raise DataFetchError("No data rows found in table")

        return pd.DataFrame(data, columns=headers)

    def fetch_sector(self, ticker: str) -> str:
        """
        Fetch sector information for a specific stock.

        Args:
            ticker: Stock ticker symbol

        Returns:
            Sector name or error message
        """
        self._rate_limit()

        url = f"{self.config.details_url}?papel={ticker}"

        try:
            response = self.session.get(url, timeout=self.config.timeout)
            response.raise_for_status()

            soup = BeautifulSoup(response.text, 'html.parser')
            table = soup.find('table', {'class': 'w728'})

            if not table:
                return "Setor não encontrado"

            # Look for sector information
            for row in table.find_all('tr'):
                cols = row.find_all('td')
                if len(cols) >= 2:
                    label = cols[0].get_text(strip=True)
                    if 'Setor' in label:
                        return cols[1].get_text(strip=True)

            return "Setor não encontrado"

        except Exception as e:
            logger.warning(f"Error fetching sector for {ticker}: {e}")
            return "Erro ao obter setor"

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.session.close()


# ============= Sector Enrichment =============
class SectorEnricher:
    """Enriches stock data with sector information."""

    def __init__(self, scraper: FundamentusScraper, cache: SectorCache, config: ScraperConfig):
        self.scraper = scraper
        self.cache = cache
        self.config = config

    def enrich_with_sectors(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Add sector information to DataFrame.

        Args:
            df: DataFrame with stock data

        Returns:
            DataFrame with sector column added
        """
        logger.info("Enriching stocks with sector information...")

        # Create a copy to avoid modifying original
        df = df.copy()

        # Initialize sector column
        df['Setor'] = 'Processando...'

        # Separate cached and uncached tickers
        cached_tickers = []
        uncached_tickers = []

        for ticker in df['Papel']:
            if self.cache.contains(ticker):
                cached_tickers.append(ticker)
            else:
                uncached_tickers.append(ticker)

        # Apply cached sectors
        for ticker in cached_tickers:
            df.loc[df['Papel'] == ticker, 'Setor'] = self.cache.get(ticker)

        logger.info(f"Applied {len(cached_tickers)} cached sectors")

        # Fetch uncached sectors in parallel
        if uncached_tickers:
            logger.info(f"Fetching {len(uncached_tickers)} sectors from web...")
            self._fetch_sectors_parallel(df, uncached_tickers)

        return df

    def _fetch_sectors_parallel(self, df: pd.DataFrame, tickers: List[str]) -> None:
        """Fetch sectors in parallel using thread pool."""
        with ThreadPoolExecutor(max_workers=self.config.max_workers) as executor:
            future_to_ticker = {
                executor.submit(self._fetch_and_cache_sector, ticker): ticker
                for ticker in tickers
            }

            for future in as_completed(future_to_ticker):
                ticker = future_to_ticker[future]
                try:
                    sector = future.result()
                    df.loc[df['Papel'] == ticker, 'Setor'] = sector
                except Exception as e:
                    logger.error(f"Error fetching sector for {ticker}: {e}")
                    df.loc[df['Papel'] == ticker, 'Setor'] = "Erro ao obter setor"

    def _fetch_and_cache_sector(self, ticker: str) -> str:
        """Fetch sector and update cache."""
        sector = self.scraper.fetch_sector(ticker)
        if sector and sector not in ["Erro ao obter setor", "Setor não encontrado"]:
            self.cache.set(ticker, sector)
        return sector


# ============= Data Processing =============
class DataProcessor:
    """Process and clean stock data."""

    @staticmethod
    def clean_percentage(value: str) -> float:
        """Convert percentage string to float."""
        if pd.isna(value) or value == '':
            return 0.0
        return float(
            value.replace('%', '')
            .replace('.', '')
            .replace(',', '.')
        ) / 100

    @staticmethod
    def clean_decimal(value: str) -> float:
        """Convert decimal string to float."""
        if pd.isna(value) or value == '':
            return 0.0
        # Handle both formats: "1.234,56" and "1234.56"
        value = str(value).replace('.', '').replace(',', '.')
        try:
            return float(value)
        except ValueError:
            return 0.0

    def clean_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Clean and convert data types in DataFrame.

        Args:
            df: Raw DataFrame from scraper

        Returns:
            Cleaned DataFrame with proper data types

        Raises:
            DataProcessingError: If data cleaning fails
        """
        try:
            df = df.copy()

            # Check for required columns
            self._validate_columns(df)

            # Find debt/equity column (might have variations)
            debt_col = self._find_debt_column(df)

            # Convert numeric columns
            df['P/L'] = df['P/L'].apply(self.clean_decimal)
            df['P/VP'] = df['P/VP'].apply(self.clean_decimal)
            df['Div.Yield'] = df['Div.Yield'].apply(self.clean_percentage)
            df['ROE'] = df['ROE'].apply(self.clean_percentage)
            df['Liq.2meses'] = df['Liq.2meses'].apply(self.clean_decimal)
            df[debt_col] = df[debt_col].apply(self.clean_decimal)
            df['Cresc. Rec.5a'] = df['Cresc. Rec.5a'].apply(self.clean_percentage)

            # Rename columns for consistency
            column_renames = {
                'Liq.2meses': 'Ltg.2meses',
                debt_col: 'Dív.Brut/Patrim',
                'Cresc. Rec.5a': 'Cresc.Rec.5a'
            }
            df = df.rename(columns=column_renames)

            # Validate data
            self._validate_data(df)

            logger.info("Data cleaning completed successfully")
            return df

        except Exception as e:
            logger.error(f"Data cleaning failed: {e}")
            raise DataProcessingError(f"Failed to clean data: {e}") from e

    def _validate_columns(self, df: pd.DataFrame) -> None:
        """Validate that required columns exist."""
        missing_columns = []
        for col in REQUIRED_COLUMNS:
            if col not in df.columns:
                # Check for variations
                found = False
                for df_col in df.columns:
                    if col.lower() in df_col.lower():
                        found = True
                        break
                if not found:
                    missing_columns.append(col)

        if missing_columns:
            raise DataProcessingError(f"Missing required columns: {missing_columns}")

    def _find_debt_column(self, df: pd.DataFrame) -> str:
        """Find the debt/equity ratio column."""
        for col in df.columns:
            if 'Dív.Brut/ Patrim.' in col or 'Dív.Brut/Patrim' in col:
                return col
        raise DataProcessingError("Debt/Equity column not found")

    def _validate_data(self, df: pd.DataFrame) -> None:
        """Validate cleaned data."""
        numeric_columns = ['P/L', 'P/VP', 'Div.Yield', 'ROE', 'Ltg.2meses',
                           'Dív.Brut/Patrim', 'Cresc.Rec.5a']

        for col in numeric_columns:
            if col in df.columns:
                nan_count = df[col].isna().sum()
                if nan_count > 0:
                    logger.warning(f"Column '{col}' has {nan_count} NaN values")

                # Check for infinite values
                inf_count = df[col].isin([float('inf'), float('-inf')]).sum()
                if inf_count > 0:
                    logger.warning(f"Column '{col}' has {inf_count} infinite values")
                    df[col].replace([float('inf'), float('-inf')], 0, inplace=True)


# ============= Filtering and Scoring =============
class StockAnalyzer:
    """Analyze and score stocks based on fundamental metrics."""

    def __init__(self, filter_criteria: FilterCriteria, score_weights: ScoreWeights):
        self.filter_criteria = filter_criteria
        self.score_weights = score_weights

    def apply_filters(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Apply filtering criteria to DataFrame.

        Args:
            df: Cleaned DataFrame

        Returns:
            Filtered DataFrame
        """
        initial_count = len(df)

        # Create filter mask
        mask = (
                (df['P/L'] > self.filter_criteria.min_pl) &
                (df['P/L'] < self.filter_criteria.max_pl) &
                (df['P/VP'] > self.filter_criteria.min_pvp) &
                (df['P/VP'] < self.filter_criteria.max_pvp) &
                (df['ROE'] > self.filter_criteria.min_roe) &
                (df['ROE'] < self.filter_criteria.max_roe) &
                (df['Div.Yield'] > self.filter_criteria.min_dividend_yield) &
                (df['Div.Yield'] < self.filter_criteria.max_dividend_yield) &
                (df['Cresc.Rec.5a'] > self.filter_criteria.min_revenue_growth) &
                (df['Dív.Brut/Patrim'] < self.filter_criteria.max_debt_equity) &
                (df['Ltg.2meses'] > self.filter_criteria.min_liquidity)
        )

        filtered_df = df[mask].copy()

        logger.info(f"Filtered {initial_count} stocks to {len(filtered_df)}")
        return filtered_df

    def calculate_score(self, row: pd.Series) -> int:
        """
        Calculate investment score for a stock.

        Args:
            row: DataFrame row with stock data

        Returns:
            Score from 0 to 14 (7 metrics × 2 points max)
        """
        score = 0

        # P/L scoring (lower is better)
        if row['P/L'] <= self.score_weights.pl_excellent:
            score += 2
        elif row['P/L'] <= self.score_weights.pl_good:
            score += 1

        # P/VP scoring (lower is better)
        if row['P/VP'] <= self.score_weights.pvp_excellent:
            score += 2
        elif row['P/VP'] <= self.score_weights.pvp_good:
            score += 1

        # Dividend Yield scoring (higher is better)
        if row['Div.Yield'] >= self.score_weights.div_yield_excellent:
            score += 2
        elif row['Div.Yield'] >= self.score_weights.div_yield_good:
            score += 1

        # ROE scoring (higher is better)
        if row['ROE'] >= self.score_weights.roe_excellent:
            score += 2
        elif row['ROE'] >= self.score_weights.roe_good:
            score += 1

        # Revenue Growth scoring (higher is better)
        if row['Cresc.Rec.5a'] >= self.score_weights.growth_excellent:
            score += 2
        elif row['Cresc.Rec.5a'] >= self.score_weights.growth_good:
            score += 1

        # Debt/Equity scoring (lower is better)
        if row['Dív.Brut/Patrim'] <= self.score_weights.debt_excellent:
            score += 2
        elif row['Dív.Brut/Patrim'] <= self.score_weights.debt_good:
            score += 1

        # Liquidity scoring (higher is better)
        if row['Ltg.2meses'] >= self.score_weights.liquidity_excellent:
            score += 2
        elif row['Ltg.2meses'] >= self.score_weights.liquidity_good:
            score += 1

        return min(score, 14)  # Maximum 14 points (7 metrics × 2)

    def add_scores(self, df: pd.DataFrame) -> pd.DataFrame:
        """Add score column to DataFrame."""
        df['Nota'] = df.apply(self.calculate_score, axis=1)
        return df

    def get_top_stocks(self, df: pd.DataFrame, n: int = 30) -> pd.DataFrame:
        """
        Get top N stocks by score.

        Args:
            df: DataFrame with scores
            n: Number of top stocks to return

        Returns:
            Top N stocks sorted by score
        """
        return df.nlargest(n, ['Nota', 'Div.Yield'])

    def get_top_by_sector(self, df: pd.DataFrame, n: int = 5) -> pd.DataFrame:
        """
        Get top N stocks from each sector.

        Args:
            df: DataFrame with scores and sectors
            n: Number of top stocks per sector

        Returns:
            DataFrame with top stocks by sector
        """
        if df.empty or 'Setor' not in df.columns:
            return pd.DataFrame()

        return (df.sort_values(['Setor', 'Nota', 'Div.Yield'],
                               ascending=[True, False, False])
                .groupby('Setor')
                .head(n)
                .sort_values(['Setor', 'Nota'], ascending=[True, False]))


# ============= Excel Export =============
class ExcelExporter:
    """Export data to formatted Excel file."""

    @staticmethod
    def save_to_excel(
            df_all: pd.DataFrame,
            df_sector: pd.DataFrame,
            output_path: Path
    ) -> None:
        """
        Save DataFrames to formatted Excel file.

        Args:
            df_all: DataFrame with all top stocks
            df_sector: DataFrame with top stocks by sector
            output_path: Path to save Excel file
        """
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Save top stocks
                df_all.to_excel(writer, index=False, sheet_name='Top 30 Ações')
                ExcelExporter._format_worksheet(
                    writer.book['Top 30 Ações'],
                    df_all
                )

                # Save top by sector if available
                if not df_sector.empty and 'Setor' in df_sector.columns:
                    df_sector.to_excel(writer, index=False, sheet_name='Top por Setor')
                    ExcelExporter._format_worksheet(
                        writer.book['Top por Setor'],
                        df_sector
                    )

            logger.info(f"Excel file saved to {output_path}")

        except Exception as e:
            logger.error(f"Failed to save Excel file: {e}")
            raise

    @staticmethod
    def _format_worksheet(worksheet: Worksheet, df: pd.DataFrame) -> None:
        """Apply formatting to Excel worksheet."""
        # Header formatting
        header_fill = PatternFill(
            start_color=ColorScheme.HEADER_COLOR.value,
            end_color=ColorScheme.HEADER_COLOR.value,
            fill_type='solid'
        )
        header_font = Font(color=ColorScheme.HEADER_TEXT.value, bold=True)
        align_center = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Format header row
        for col_idx in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center
            cell.border = thin_border

            # Adjust column width
            column_letter = get_column_letter(col_idx)
            col_name = df.columns[col_idx - 1]

            if col_name == 'Setor':
                worksheet.column_dimensions[column_letter].width = 28
            elif col_name == 'Papel':
                worksheet.column_dimensions[column_letter].width = 10
            else:
                worksheet.column_dimensions[column_letter].width = 12

        # Format data rows
        for row_idx in range(2, len(df) + 2):
            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

        # Apply number formatting
        ExcelExporter._apply_number_formats(worksheet, df)

        # Apply conditional formatting to score column
        if 'Nota' in df.columns:
            ExcelExporter._apply_score_formatting(worksheet, df)

        # Freeze header row
        worksheet.freeze_panes = 'A2'

    @staticmethod
    def _apply_number_formats(worksheet: Worksheet, df: pd.DataFrame) -> None:
        """Apply number formatting to specific columns."""
        format_mappings = {
            'Div.Yield': '0.00%',
            'ROE': '0.00%',
            'Cresc.Rec.5a': '0.00%',
            'P/L': '0.00',
            'P/VP': '0.00',
            'Dív.Brut/Patrim': '0.00',
            'Ltg.2meses': '#,##0'
        }

        for col_name, format_str in format_mappings.items():
            if col_name in df.columns:
                col_idx = df.columns.get_loc(col_name) + 1
                for row_idx in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.number_format = format_str

    @staticmethod
    def _apply_score_formatting(worksheet: Worksheet, df: pd.DataFrame) -> None:
        """Apply conditional formatting to score column."""
        nota_col_idx = df.columns.get_loc('Nota') + 1

        for row_idx in range(2, len(df) + 2):
            cell = worksheet.cell(row=row_idx, column=nota_col_idx)

            if cell.value is not None:
                if cell.value >= 10:
                    fill_color = ColorScheme.HIGH_SCORE.value
                elif cell.value >= 7:
                    fill_color = ColorScheme.MEDIUM_SCORE.value
                else:
                    fill_color = ColorScheme.LOW_SCORE.value

                cell.fill = PatternFill(
                    start_color=fill_color,
                    end_color=fill_color,
                    fill_type='solid'
                )


# ============= Report Generator =============
class ReportGenerator:
    """Generate analysis reports."""

    @staticmethod
    def print_summary(df_all: pd.DataFrame, df_sector: pd.DataFrame) -> None:
        """Print summary of analysis results."""
        print(f"\n{'=' * 60}")
        print("  📈 ANÁLISE FUNDAMENTALISTA DE AÇÕES - RESULTADO")
        print(f"{'=' * 60}")
        print(f"✅ Total de ações selecionadas: {len(df_all)}")

        # Print basic statistics
        if not df_all.empty:
            print(f"\n📊 ESTATÍSTICAS DO PORTFÓLIO:")
            print(f"  • Nota média: {df_all['Nota'].mean():.2f}")
            print(f"  • Dividend Yield médio: {df_all['Div.Yield'].mean():.2%}")
            print(f"  • P/L médio: {df_all['P/L'].mean():.2f}")
            print(f"  • ROE médio: {df_all['ROE'].mean():.2%}")

        # Print top stocks by sector
        if not df_sector.empty and 'Setor' in df_sector.columns:
            print(f"\n🏆 TOP AÇÕES POR SETOR:")
            print("-" * 60)

            for sector, group in df_sector.groupby('Setor'):
                if sector and sector not in ["Setor não encontrado", "Erro ao obter setor"]:
                    print(f"\n📌 {sector}")
                    print("-" * 40)

                    display_df = group[['Papel', 'Nota', 'Div.Yield', 'P/L', 'ROE']].copy()
                    display_df['Div.Yield'] = display_df['Div.Yield'].apply(lambda x: f"{x:.2%}")
                    display_df['ROE'] = display_df['ROE'].apply(lambda x: f"{x:.2%}")
                    display_df['P/L'] = display_df['P/L'].apply(lambda x: f"{x:.2f}")

                    print(display_df.to_string(index=False))

        print(f"\n{'=' * 60}")


# ============= Main Application =============
class StockApplication:
    """Main application for stock analysis."""

    def __init__(
            self,
            config: Optional[ScraperConfig] = None,
            filter_criteria: Optional[FilterCriteria] = None,
            score_weights: Optional[ScoreWeights] = None
    ):
        self.config = config or ScraperConfig()
        self.filter_criteria = filter_criteria or FilterCriteria()
        self.score_weights = score_weights or ScoreWeights()
        self.processor = DataProcessor()
        self.analyzer = StockAnalyzer(self.filter_criteria, self.score_weights)

    def check_existing_file(self, filepath: Path) -> bool:
        """Check if output file exists and handle user confirmation."""
        if filepath.exists():
            logger.warning(f"File '{filepath}' already exists")
            # In production, you might want to handle this differently
            return True
        return True

    def run(self) -> None:
        """Run the complete stock analysis pipeline."""
        output_path = Path(self.config.output_filename)
        cache_path = Path(self.config.cache_filename)

        print("\n" + "=" * 60)
        print("  📊 ANÁLISE DE AÇÕES - FUNDAMENTUS")
        print("=" * 60)
        print(f"📍 Fonte: {self.config.base_url}")
        print(f"💾 Cache: {cache_path}")
        print("=" * 60)

        if not self.check_existing_file(output_path):
            print("Operação cancelada.")
            return

        try:
            # Initialize cache
            with SectorCache(cache_path) as cache:
                print(f"\n📂 Cache carregado: {cache.size} setores")

                # Initialize scraper
                with FundamentusScraper(self.config) as scraper:
                    # Fetch data
                    print("\n⏳ Coletando dados do Fundamentus...")
                    raw_df = scraper.fetch_stock_list()

                    # Process data
                    print("🔧 Processando dados...")
                    clean_df = self.processor.clean_data(raw_df)

                    # Apply filters
                    print("🔍 Aplicando filtros fundamentalistas...")
                    filtered_df = self.analyzer.apply_filters(clean_df)

                    if filtered_df.empty:
                        print("\n⚠️  Nenhuma ação atende aos critérios especificados.")
                        return

                    # Calculate scores
                    print("📊 Calculando pontuações...")
                    scored_df = self.analyzer.add_scores(filtered_df)

                    # Get top stocks
                    print(f"🎯 Selecionando top {self.config.top_stocks_limit} ações...")
                    top_stocks = self.analyzer.get_top_stocks(
                        scored_df,
                        self.config.top_stocks_limit
                    )

                    # Enrich with sectors
                    print("\n🏢 Identificando setores...")
                    enricher = SectorEnricher(scraper, cache, self.config)
                    top_stocks = enricher.enrich_with_sectors(top_stocks)

                    # Select final columns
                    columns_to_keep = [col for col in FINAL_COLUMNS if col in top_stocks.columns]
                    final_df = top_stocks[columns_to_keep]

                    # Get top by sector
                    top_by_sector = self.analyzer.get_top_by_sector(
                        final_df,
                        self.config.top_per_sector
                    )

                    # Save to Excel
                    print("\n💾 Salvando resultados...")
                    ExcelExporter.save_to_excel(final_df, top_by_sector, output_path)

                    # Print summary
                    ReportGenerator.print_summary(final_df, top_by_sector)

                    print(f"\n✅ Arquivo salvo em: {output_path.absolute()}")
                    print(f"💾 Cache atualizado: {cache.size} setores")

        except StockScraperError as e:
            logger.error(f"Application error: {e}")
            print(f"\n❌ Erro: {e}")
        except Exception as e:
            logger.error(f"Unexpected error: {e}", exc_info=True)
            print(f"\n❌ Erro inesperado: {e}")


def main():
    """Entry point for the application."""
    app = StockApplication()
    app.run()


if __name__ == "__main__":
    main()