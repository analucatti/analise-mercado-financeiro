"""
FII (Fundos de Investimento Imobiliário) Scraper and Analyzer
Scrapes real estate investment fund data from Fundamentus and analyzes it.
"""

import logging
from dataclasses import dataclass
from enum import Enum
from pathlib import Path
from typing import Optional

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


# ============= Configuration =============
@dataclass
class ScraperConfig:
    """Configuration for the FII scraper."""
    base_url: str = "https://www.fundamentus.com.br/fii_resultado.php"
    timeout: int = 10
    user_agent: str = (
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 '
        '(KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
    )
    output_filename: str = "fundos_imobiliarios_filtrados.xlsx"
    top_n_per_segment: int = 5


@dataclass
class FilterCriteria:
    """Criteria for filtering FII funds."""
    min_dividend_yield: float = 0.07  # 7%
    max_dividend_yield: float = 0.25  # 25%
    min_pvp: float = 0.5
    max_pvp: float = 1.1
    min_liquidity: float = 1_000_000
    min_market_value: float = 1_000_000_000


@dataclass
class ScoreWeights:
    """Scoring weights and thresholds for FII evaluation."""
    dividend_yield_high: float = 0.14
    dividend_yield_medium: float = 0.12
    pvp_low: float = 0.80
    pvp_medium: float = 0.85
    liquidity_high: float = 5_000_000
    liquidity_medium: float = 2_000_000
    market_value_high: float = 2_000_000_000
    market_value_medium: float = 1_500_000_000
    vacancy_low: float = 0.05
    vacancy_medium: float = 0.10


class ColorScheme(Enum):
    """Excel color schemes for formatting."""
    HEADER_COLOR = '4F81BD'
    HEADER_TEXT = 'FFFFFF'
    HIGH_SCORE = 'C6EFCE'
    MEDIUM_SCORE = 'FFEB9C'
    LOW_SCORE = 'FFC7CE'


# ============= Column Mappings =============
COLUMN_MAPPINGS = {
    'Papel': 'Papel',
    'Segmento': 'Segmento',
    'Cotação': 'Cotacao',
    'FFO Yield': 'FFO Yield',
    'Dividend Yield': 'Dividend Yield',
    'P/VP': 'P/VP',
    'Valor de Mercado': 'Valor de Mercado',
    'Liquidez': 'Liquidez',
    'Qtd de imóveis': 'Qtd Imoveis',
    'Preço do m2': 'Preco m2',
    'Aluguel por m2': 'Aluguel m2',
    'Cap Rate': 'Cap Rate',
    'Vacância Média': 'Vacancia Media'
}

FINAL_COLUMNS = [
    'Papel', 'Segmento', 'Dividend Yield', 'P/VP',
    'Valor de Mercado', 'Liquidez', 'Qtd Imoveis',
    'Vacancia Media', 'Nota'
]


# ============= Exceptions =============
class FIIScraperError(Exception):
    """Base exception for FII scraper."""
    pass


class DataFetchError(FIIScraperError):
    """Exception raised when data fetching fails."""
    pass


class DataProcessingError(FIIScraperError):
    """Exception raised when data processing fails."""
    pass


# ============= Data Fetching =============
class FundamentusScraper:
    """Scraper for Fundamentus FII data."""

    def __init__(self, config: ScraperConfig):
        self.config = config
        self.session = self._create_session()

    def _create_session(self) -> requests.Session:
        """Create a configured requests session."""
        session = requests.Session()
        session.headers.update({'User-Agent': self.config.user_agent})
        return session

    def fetch_data(self) -> pd.DataFrame:
        """
        Fetch FII data from Fundamentus website.

        Returns:
            DataFrame with raw FII data

        Raises:
            DataFetchError: If data cannot be fetched or parsed
        """
        try:
            logger.info(f"Fetching data from {self.config.base_url}")
            response = self.session.get(
                self.config.base_url,
                timeout=self.config.timeout
            )
            response.raise_for_status()

            soup = BeautifulSoup(response.text, 'html.parser')
            table = soup.find('table', {'id': 'tabelaResultado'})

            if not table:
                raise DataFetchError("Table 'tabelaResultado' not found on page")

            df = self._parse_table(table)
            logger.info(f"Successfully fetched {len(df)} records")
            return df

        except requests.RequestException as e:
            logger.error(f"Request failed: {e}")
            raise DataFetchError(f"Failed to fetch data: {e}") from e
        except Exception as e:
            logger.error(f"Unexpected error during data fetch: {e}")
            raise DataFetchError(f"Unexpected error: {e}") from e

    def _parse_table(self, table) -> pd.DataFrame:
        """Parse HTML table into DataFrame."""
        headers = [th.get_text(strip=True) for th in table.find('tr').find_all('th')]

        data = []
        for row in table.find_all('tr')[1:]:
            row_data = [td.get_text(strip=True) for td in row.find_all('td')]
            if row_data:  # Skip empty rows
                data.append(row_data)

        if not data:
            raise DataFetchError("No data rows found in table")

        df = pd.DataFrame(data, columns=headers)
        df = df.rename(columns=COLUMN_MAPPINGS)

        return df

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.session.close()


# ============= Data Processing =============
class DataProcessor:
    """Process and clean FII data."""

    @staticmethod
    def clean_percentage(value: str) -> float:
        """Convert percentage string to float."""
        if pd.isna(value) or value == '':
            return 0.0
        return float(
            value.replace('%', '')
            .replace('.', '')
            .replace(',', '.')
        ) / 100

    @staticmethod
    def clean_decimal(value: str) -> float:
        """Convert decimal string to float."""
        if pd.isna(value) or value == '':
            return 0.0
        return float(value.replace(',', '.'))

    @staticmethod
    def clean_integer(value: str) -> float:
        """Convert integer string with dots as thousand separators to float."""
        if pd.isna(value) or value == '':
            return 0.0
        if isinstance(value, str):
            return float(value.replace('.', ''))
        return float(value)

    def clean_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Clean and convert data types in DataFrame.

        Args:
            df: Raw DataFrame from scraper

        Returns:
            Cleaned DataFrame with proper data types

        Raises:
            DataProcessingError: If data cleaning fails
        """
        try:
            df = df.copy()

            # Convert percentage columns
            df['Dividend Yield'] = df['Dividend Yield'].apply(self.clean_percentage)
            df['Vacancia Media'] = df['Vacancia Media'].apply(self.clean_percentage)

            # Convert decimal columns
            df['P/VP'] = df['P/VP'].apply(self.clean_decimal)

            # Convert integer columns
            df['Valor de Mercado'] = df['Valor de Mercado'].apply(self.clean_integer)
            df['Liquidez'] = df['Liquidez'].apply(self.clean_integer)

            # Validate data
            self._validate_data(df)

            logger.info("Data cleaning completed successfully")
            return df

        except Exception as e:
            logger.error(f"Data cleaning failed: {e}")
            raise DataProcessingError(f"Failed to clean data: {e}") from e

    def _validate_data(self, df: pd.DataFrame) -> None:
        """Validate cleaned data."""
        required_columns = ['Dividend Yield', 'P/VP', 'Valor de Mercado', 'Liquidez']

        for col in required_columns:
            if col not in df.columns:
                raise DataProcessingError(f"Required column '{col}' not found")

            # Check for NaN values in critical columns
            nan_count = df[col].isna().sum()
            if nan_count > 0:
                logger.warning(f"Column '{col}' has {nan_count} NaN values")


# ============= Filtering and Scoring =============
class FIIAnalyzer:
    """Analyze and score FII funds."""

    def __init__(self, filter_criteria: FilterCriteria, score_weights: ScoreWeights):
        self.filter_criteria = filter_criteria
        self.score_weights = score_weights

    def apply_filters(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Apply filtering criteria to DataFrame.

        Args:
            df: Cleaned DataFrame

        Returns:
            Filtered DataFrame
        """
        initial_count = len(df)

        filtered_df = df[
            (df['Dividend Yield'] > self.filter_criteria.min_dividend_yield) &
            (df['Dividend Yield'] < self.filter_criteria.max_dividend_yield) &
            (df['P/VP'] > self.filter_criteria.min_pvp) &
            (df['P/VP'] < self.filter_criteria.max_pvp) &
            (df['Liquidez'] > self.filter_criteria.min_liquidity) &
            (df['Valor de Mercado'] > self.filter_criteria.min_market_value)
            ].copy()

        logger.info(f"Filtered {initial_count} records to {len(filtered_df)}")
        return filtered_df

    def calculate_score(self, row: pd.Series) -> int:
        """
        Calculate investment score for a FII.

        Args:
            row: DataFrame row with FII data

        Returns:
            Score from 0 to 10
        """
        score = 0

        # Dividend Yield scoring
        if row['Dividend Yield'] >= self.score_weights.dividend_yield_high:
            score += 2
        elif row['Dividend Yield'] >= self.score_weights.dividend_yield_medium:
            score += 1

        # P/VP scoring (lower is better)
        if row['P/VP'] <= self.score_weights.pvp_low:
            score += 2
        elif row['P/VP'] <= self.score_weights.pvp_medium:
            score += 1

        # Liquidity scoring
        if row['Liquidez'] >= self.score_weights.liquidity_high:
            score += 2
        elif row['Liquidez'] >= self.score_weights.liquidity_medium:
            score += 1

        # Market Value scoring
        if row['Valor de Mercado'] >= self.score_weights.market_value_high:
            score += 2
        elif row['Valor de Mercado'] >= self.score_weights.market_value_medium:
            score += 1

        # Vacancy scoring (lower is better)
        if pd.notna(row.get('Vacancia Media')):
            if row['Vacancia Media'] <= self.score_weights.vacancy_low:
                score += 2
            elif row['Vacancia Media'] <= self.score_weights.vacancy_medium:
                score += 1

        return min(score, 10)

    def add_scores(self, df: pd.DataFrame) -> pd.DataFrame:
        """Add score column to DataFrame."""
        df['Nota'] = df.apply(self.calculate_score, axis=1)
        return df

    def get_top_by_segment(self, df: pd.DataFrame, n: int = 5) -> pd.DataFrame:
        """
        Get top N funds from each segment.

        Args:
            df: DataFrame with scores
            n: Number of top funds per segment

        Returns:
            DataFrame with top funds by segment
        """
        if df.empty:
            return pd.DataFrame()

        return (df.sort_values(['Segmento', 'Nota', 'Dividend Yield'],
                               ascending=[True, False, False])
                .groupby('Segmento')
                .head(n)
                .sort_values(['Segmento', 'Nota'], ascending=[True, False]))


# ============= Excel Export =============
class ExcelExporter:
    """Export data to formatted Excel file."""

    @staticmethod
    def save_to_excel(
            df_all: pd.DataFrame,
            df_top: pd.DataFrame,
            output_path: Path
    ) -> None:
        """
        Save DataFrames to formatted Excel file.

        Args:
            df_all: DataFrame with all filtered funds
            df_top: DataFrame with top funds by segment
            output_path: Path to save Excel file
        """
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Save all funds
                df_all.to_excel(writer, index=False, sheet_name='Todos Fundos')
                ExcelExporter._format_worksheet(
                    writer.book['Todos Fundos'],
                    df_all
                )

                # Save top funds by segment
                if not df_top.empty:
                    df_top.to_excel(writer, index=False, sheet_name='Top por Segmento')
                    ExcelExporter._format_worksheet(
                        writer.book['Top por Segmento'],
                        df_top
                    )

            logger.info(f"Excel file saved to {output_path}")

        except Exception as e:
            logger.error(f"Failed to save Excel file: {e}")
            raise

    @staticmethod
    def _format_worksheet(worksheet: Worksheet, df: pd.DataFrame) -> None:
        """Apply formatting to Excel worksheet."""
        # Header formatting
        header_fill = PatternFill(
            start_color=ColorScheme.HEADER_COLOR.value,
            end_color=ColorScheme.HEADER_COLOR.value,
            fill_type='solid'
        )
        header_font = Font(color=ColorScheme.HEADER_TEXT.value, bold=True)
        align_center = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Format header row
        for col_idx in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center
            cell.border = thin_border

            # Adjust column width
            column_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[column_letter].width = 17

        # Format data rows
        for row_idx in range(2, len(df) + 2):
            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

        # Apply number formatting
        ExcelExporter._apply_number_formats(worksheet, df)

        # Apply conditional formatting to score column
        if 'Nota' in df.columns:
            ExcelExporter._apply_score_formatting(worksheet, df)

        # Freeze header row
        worksheet.freeze_panes = 'A2'

    @staticmethod
    def _apply_number_formats(worksheet: Worksheet, df: pd.DataFrame) -> None:
        """Apply number formatting to specific columns."""
        format_mappings = {
            'Dividend Yield': '0.00%',
            'Vacancia Media': '0.00%',
            'P/VP': '0.00',
            'Valor de Mercado': '#,##0',
            'Liquidez': '#,##0',
            'Qtd Imoveis': '0'
        }

        for col_name, format_str in format_mappings.items():
            if col_name in df.columns:
                col_idx = df.columns.get_loc(col_name) + 1
                for row_idx in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.number_format = format_str

    @staticmethod
    def _apply_score_formatting(worksheet: Worksheet, df: pd.DataFrame) -> None:
        """Apply conditional formatting to score column."""
        nota_col_idx = df.columns.get_loc('Nota') + 1

        for row_idx in range(2, len(df) + 2):
            cell = worksheet.cell(row=row_idx, column=nota_col_idx)

            if cell.value is not None:
                if cell.value >= 8:
                    fill_color = ColorScheme.HIGH_SCORE.value
                elif cell.value >= 5:
                    fill_color = ColorScheme.MEDIUM_SCORE.value
                else:
                    fill_color = ColorScheme.LOW_SCORE.value

                cell.fill = PatternFill(
                    start_color=fill_color,
                    end_color=fill_color,
                    fill_type='solid'
                )


# ============= Report Generator =============
class ReportGenerator:
    """Generate analysis reports."""

    @staticmethod
    def print_summary(df_all: pd.DataFrame, df_top: pd.DataFrame) -> None:
        """Print summary of analysis results."""
        print(f"\n{'=' * 50}")
        print("ANÁLISE CONCLUÍDA COM SUCESSO")
        print(f"{'=' * 50}")
        print(f"Total de fundos filtrados: {len(df_all)}")

        if not df_top.empty:
            print("\n📊 TOP 5 FUNDOS POR SEGMENTO:")
            print("-" * 50)

            for segment, group in df_top.groupby('Segmento'):
                print(f"\n🏢 Segmento: {segment}")
                print("-" * 30)

                display_df = group[['Papel', 'Nota', 'Dividend Yield', 'P/VP']].copy()
                display_df['Dividend Yield'] = display_df['Dividend Yield'].apply(
                    lambda x: f"{x:.2%}"
                )
                display_df['P/VP'] = display_df['P/VP'].apply(lambda x: f"{x:.2f}")

                print(display_df.to_string(index=False))

        print(f"\n{'=' * 50}")


# ============= Main Application =============
class FIIApplication:
    """Main application for FII analysis."""

    def __init__(
            self,
            config: Optional[ScraperConfig] = None,
            filter_criteria: Optional[FilterCriteria] = None,
            score_weights: Optional[ScoreWeights] = None
    ):
        self.config = config or ScraperConfig()
        self.filter_criteria = filter_criteria or FilterCriteria()
        self.score_weights = score_weights or ScoreWeights()
        self.processor = DataProcessor()
        self.analyzer = FIIAnalyzer(self.filter_criteria, self.score_weights)

    def check_existing_file(self, filepath: Path) -> bool:
        """Check if output file exists and handle user confirmation."""
        if filepath.exists():
            logger.warning(f"File '{filepath}' already exists")
            # In production, you might want to handle this differently
            # For now, we'll overwrite
            return True
        return True

    def run(self) -> None:
        """Run the complete FII analysis pipeline."""
        output_path = Path(self.config.output_filename)

        print("\n" + "=" * 60)
        print("  FILTRO DE FUNDOS IMOBILIÁRIOS - ANÁLISE FUNDAMENTALISTA")
        print("=" * 60)
        print(f"📍 Fonte: {self.config.base_url}")
        print("=" * 60)

        if not self.check_existing_file(output_path):
            print("Operação cancelada.")
            return

        try:
            # Fetch data
            print("\n⏳ Iniciando coleta de dados...")
            with FundamentusScraper(self.config) as scraper:
                raw_df = scraper.fetch_data()

            # Process data
            print("🔧 Processando dados...")
            clean_df = self.processor.clean_data(raw_df)

            # Apply filters
            print("🔍 Aplicando filtros...")
            filtered_df = self.analyzer.apply_filters(clean_df)

            if filtered_df.empty:
                print("\n⚠️  Nenhum fundo atende aos critérios especificados.")
                return

            # Calculate scores
            print("📊 Calculando pontuações...")
            scored_df = self.analyzer.add_scores(filtered_df)

            # Prepare final DataFrames
            final_df = scored_df[FINAL_COLUMNS].sort_values(
                by=['Nota', 'Dividend Yield'],
                ascending=[False, False]
            )

            top_df = self.analyzer.get_top_by_segment(
                final_df,
                self.config.top_n_per_segment
            )

            # Save to Excel
            print("💾 Salvando resultados...")
            ExcelExporter.save_to_excel(final_df, top_df, output_path)

            # Print summary
            ReportGenerator.print_summary(final_df, top_df)

            print(f"\n✅ Arquivo salvo em: {output_path.absolute()}")

        except FIIScraperError as e:
            logger.error(f"Application error: {e}")
            print(f"\n❌ Erro: {e}")
        except Exception as e:
            logger.error(f"Unexpected error: {e}", exc_info=True)
            print(f"\n❌ Erro inesperado: {e}")


def main():
    """Entry point for the application."""
    app = FIIApplication()
    app.run()


if __name__ == "__main__":
    main()